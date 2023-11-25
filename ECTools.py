import tkinter as tk
from tkinter import filedialog
from tkinter.simpledialog import *
from tkinter.messagebox import showinfo
import openpyxl as xl
import getpass
import Libraries.quip as quip


class BaseFrame(tk.Frame):
    def __init__(self, master, controller):
        tk.Frame.__init__(self, master)
        self.controller = controller
        self.grid()
        self.create_widgets()

    def create_widgets(self):
        raise NotImplementedError


class CollectSerialsFrame(BaseFrame):
    def create_widgets(self):
        self.open_collection_btn = tk.Button(
            self,
            anchor=tk.W,
            command=self.open_collection,
            padx=5,
            pady=5,
            text="Open serial collection",
        )

        self.lbl = tk.Label(
            self, text="xxxxxx", fg="white", height=10, width=20, font="System 18 bold"
        )
        self.lbl.grid(row=0, column=0, sticky="ew", columnspan=2)
        self.txt = tk.Entry(self)
        self.txt.focus_set()
        self.txt.grid(row=1, column=0, sticky="ew", columnspan=2)
        self.txt.bind("<Return>", self.submit)
        self.open_collection_btn.grid(
            row=2, column=0, padx=5, pady=5, sticky="ew"
        )
        self.grid_columnconfigure(0, weight=1)

    def submit(self, event):
        s_num = self.txt.get()
        if self.check_in_list(s_num, self.serials):
            self.lbl.configure(bg="red", text=s_num + " is already in list.")
        else:
            self.lbl.configure(bg="green", text=s_num + " added.")

    def check_in_list(self, serial, list):
        if serial in list:
            return True
        else:
            if serial.lower() in list:
                return True
            else:
                return False

    def open_collection(self):
        self.file = tk.filedialog.askopenfilename(
            initialdir="./",
            title="Select a File",
            filetypes=(("Excel files", "*.xlsx*"), ("all files", "*.*")),
        )
        self.col_wb = xl.load_workbook(filename=self.file)
        self.col_ws = self.col_wb.active
        self.serials = []
        for row in self.col_ws.values:
            self.serials.append(row[0])


class SubtractExcelSerials(BaseFrame):
    def create_widgets(self):
        self.file_a=[]
        self.file_b=[]

        self.load_file_a_btn = tk.Button(
            self,
            anchor=tk.W,
            command=lambda: self.load_file(self.file_a),
            padx=5,
            pady=5,
            text="Load Excel Sheet A",
        )
        self.load_file_b_btn = tk.Button(
            self,
            anchor=tk.W,
            command=lambda: self.load_file(self.file_b),
            padx=5,
            pady=5,
            text="Load Excel Sheet B",
        )
        self.save_output_btn = tk.Button(
            self,
            anchor=tk.W,
            command=lambda: self.subtract_sheets(),
            padx=5,
            pady=5,
            text="Output",
        )
        self.lbl = tk.Label(
            self, text="Take serials from excel sheet B\nand remove them from excel sheet A", bg="black", fg="white", height=3, width=20, font="System 12 bold"
        )
        self.lbl.grid(row=0, column=0, sticky="ew", columnspan=2)
        self.load_file_a_btn.grid(
            row=1, column=0, padx=5, pady=5, sticky="ew"
        )
        self.load_file_b_btn.grid(
            row=1, column=1, padx=5, pady=5, sticky="ew"
        )
        self.save_output_btn.grid(
            row=2, column=0, padx=5, pady=5, sticky="ew", columnspan=2
        )
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

    def load_file(self, list):
        self.file = tk.filedialog.askopenfilename(
            initialdir="./",
            title="Select a File",
            filetypes=(("Excel files", "*.xlsx*"), ("all files", "*.*")),
        )
        self.col_wb = xl.load_workbook(filename=self.file)
        self.col_ws = self.col_wb.active
        for row in self.col_ws.values:
            list.append(row[0])

    def subtract_sheets(self):
        for i in self.file_b:
            if i in self.file_a:
                self.file_a.remove(i)
        self.controller.save_file(self.file_a)


class CompareSerialsFrame(BaseFrame):
    def create_widgets(self):
        self.list_not_scanned_btn = tk.Button(
            self,
            anchor=tk.W,
            command=lambda: self.controller.save_file(LIST_NOT_SCANNED_OUTPUT),
            padx=5,
            pady=5,
            text="Output serials in list not scanned",
        )
        self.scanned_not_list_btn = tk.Button(
            self,
            anchor=tk.E,
            command=lambda: self.controller.save_file(SCANNED_NOT_LIST_OUTPUT),
            padx=5,
            pady=5,
            text="Output serials scanned not in list",
        )
        """self.load_file_data = tk.Button(
            self,
            anchor=tk.N,
            command=lambda: self.controller.load_file(),
            padx=5,
            pady=5,
            text="Load an Excel file",
        )"""
        self.lbl = tk.Label(
            self, text="xxxxxx", fg="white", height=10, width=20, font="System 18 bold"
        )
        self.lbl.grid(row=0, column=0, sticky="ew", columnspan=2)
        self.txt = tk.Entry(self)
        self.txt.focus_set()
        self.txt.grid(row=1, column=0, sticky="ew", columnspan=2)
        self.txt.bind("<Return>", self.submit)
        self.list_not_scanned_btn.grid(
            row=3, column=0, padx=5, pady=5, sticky="ew"
        )
        self.scanned_not_list_btn.grid(
            row=3, column=1, padx=5, pady=5, sticky="ew"
        )
        """self.load_file_data.grid(
            row=4, column=0, padx=5, pady=5, sticky="ew", columnspan=2
        )"""
        # self.grid_columnconfigure(0, weight=1)

    def submit(self, event):
        s_num = self.txt.get()

        if self.check_in_list(s_num, SERIALS):
            self.lbl.configure(bg="green", text=s_num)
            if s_num in LIST_NOT_SCANNED_OUTPUT:
                try:
                    LIST_NOT_SCANNED_OUTPUT.remove(s_num)
                except ValueError:
                    LIST_NOT_SCANNED_OUTPUT.remove(s_num.lower())
            print(s_num + " = true")
        else:
            self.lbl.configure(bg="red", text=s_num)
            SCANNED_NOT_LIST_OUTPUT.append(s_num)
            print(s_num + " = false")
            print("----------------")
            for k in SERIALS:
                print(k)

        if self.check_in_list(s_num, MISSING_SERIALS):
            self.lbl.configure(bg="blue", text=s_num)
            showinfo("Missing Item", "{} is a missing item.".format(s_num))
            print(s_num + " = MISSING ITEM")

        self.txt.delete(0, tk.END)

    def check_in_list(self, serial, list):
        if serial in list:
            return True
        else:
            if serial.lower() in list:
                return True
            else:
                return False


class LookupSerialsFrame(BaseFrame):
    def config(self):
        self.serial_column = 0
        self.locator_column = 1

    def create_widgets(self):
        self.lbl = tk.Label(
            self, text="xxxxxx", fg="white", height=10, width=20, font="System 18 bold"
        )
        self.lbl.grid(row=0, column=0, sticky="ew", columnspan=2)
        self.txt = tk.Entry(self)
        self.txt.focus_set()
        self.txt.grid(row=1, column=0, sticky="ew", columnspan=2)
        self.txt.bind("<Return>", self.submit)
        self.grid_columnconfigure(0, weight=1)

    def submit(self, event):
        s_num = self.txt.get()

        if self.check_in_list(s_num, MASTER_SERIALS):
            self.lbl.configure(bg="green", text=s_num + "\n" + MASTER_SERIALS[s_num])
            print(s_num + " = true")
        else:
            self.lbl.configure(bg="red", text=s_num + " is not virtually at SAV4")
            print(s_num + " = false")

        if self.check_in_list(s_num, MISSING_SERIALS):
            self.lbl.configure(bg="blue", text=s_num)
            showinfo("Missing Item", "{} is a missing item.".format(s_num))
            print(s_num + " = MISSING ITEM")

        self.txt.delete(0, tk.END)

    def check_in_list(self, serial, list):
        if serial in list:
            return True
        else:
            if serial.lower() in list:
                return True
            else:
                return False


class HomeFrame(BaseFrame):
    def create_widgets(self):
        self.compare_scanned_serials = tk.Button(
            self,
            anchor=tk.N,
            command=lambda: self.controller.show_frame(CompareSerialsFrame),
            padx=5,
            pady=5,
            text="Compare Scanned Serials",
        )

        self.collect_serials = tk.Button(
            self,
            anchor=tk.N,
            command=lambda: self.controller.show_frame(CollectSerialsFrame),
            padx=5,
            pady=5,
            text="Collect Serials",
        )

        self.compare_excel_serials = tk.Button(
            self,
            anchor=tk.N,
            command=lambda: self.controller.show_frame(SubtractExcelSerials),
            padx=5,
            pady=5,
            text="Subtract Serials from Excel Sheet",
        )

        self.lookup_serials = tk.Button(
            self,
            anchor=tk.N,
            command=lambda: self.controller.show_frame(LookupSerialsFrame),
            padx=5,
            pady=5,
            text="Lookup Serials in Master List",
        )

        """self.load_file_data = tk.Button(
            self,
            anchor=tk.N,
            command=lambda: self.controller.load_file(),
            padx=5,
            pady=5,
            text="Load a specific Excel file for Compare",
        )"""

        self.grid_columnconfigure(0, weight=1)

        self.compare_scanned_serials.grid(row=0, column=0, sticky="ew")
        self.compare_excel_serials.grid(row=1, column=0, sticky="ew")
        self.collect_serials.grid(row=2, column=0, sticky="ew")
        self.lookup_serials.grid(row=3, column=0, sticky="ew")
        """self.load_file_data.grid(row=4, column=0, sticky="ew")"""


class EC_Toolkit(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        self.load_data_globally()
        self.title("EC Toolkit")
        self.create_widgets()
        self.resizable(0, 0)
        self.grid()
        self.grid_columnconfigure(0, weight=1)

        print(getpass.getuser())

    def create_widgets(self):
        #   Frame Container
        self.container = tk.Frame(self)
        self.container.grid(row=0, column=0, sticky="nsew")
        self.container.grid_columnconfigure(0, weight=1)

        #   Frames
        self.frames = {}
        for f in (
            HomeFrame,
            CompareSerialsFrame,
            CollectSerialsFrame,
            LookupSerialsFrame,
            SubtractExcelSerials,
        ):  # defined subclasses of BaseFrame
            frame = f(self.container, self)
            frame.grid(row=0, column=0, sticky="nsew")
            self.frames[f] = frame
        self.show_frame(HomeFrame)

    def show_frame(self, cls):
        self.frames[cls].tkraise()

    def load_data_globally(self):
        # Declare global variables
        global SERIALS
        global MASTER_SERIALS
        global MISSING_SERIALS
        global SCANNED_NOT_LIST_OUTPUT
        global LIST_NOT_SCANNED_OUTPUT

        # Load data seet to compare against

        DATA_WB = xl.load_workbook(filename="data.xlsx")
        DATA_WS = DATA_WB.active
        SERIALS = []
        self.serial_list = []

        for row in DATA_WS.values:
            SERIALS.append(row[0])
            self.serial_list.append(row[0])

        ML_WB = xl.load_workbook(filename="Data/masterlist.xlsx")
        ML_WS = ML_WB.active
        MASTER_SERIALS = {}

        for row in ML_WS.values:
            MASTER_SERIALS.update({row[0]: row[1]})

        # Load missing serials
        MISSING_WB = xl.load_workbook(filename="Data/missing.xlsx")
        MISSING_WS = MISSING_WB.active
        MISSING_SERIALS = []

        for row in MISSING_WS.values:
            MISSING_SERIALS.append(row[0])

        SCANNED_NOT_LIST_OUTPUT = []
        LIST_NOT_SCANNED_OUTPUT = SERIALS.copy()

    def load_file(self):
        # handle loading a specific excel file of serials
        self.file = tk.filedialog.askopenfilename(
            initialdir="./",
            title="Select a File",
            filetypes=(("Excel files", "*.xlsx*"), ("all files", "*.*")),
        )
        self.serial_column = askinteger(
            "Input",
            "Which column contains the serial numbers? (0-Indexed)",
            parent=self,
            minvalue=0,
            initialvalue=0,
        )
        DATA_WB = xl.load_workbook(filename=self.file)
        DATA_WS = DATA_WB.active
        self.serial_list = []

        for row in DATA_WS.values:
            SERIALS.append(row[self.serial_column])
        
        LIST_NOT_SCANNED_OUTPUT = SERIALS.copy()

    def save_file(self, list):
        self.output_wb = xl.Workbook()
        self.output_wb.create_sheet()
        self.output_ws = self.output_wb.active

        for sn in list:
            row = [sn]
            self.output_ws.append(row)
        self.output_wb.save(
            filename=filedialog.asksaveasfilename(
                initialfile="Output.xlsx", defaultextension=".xlsx"
            )
        )


if __name__ == "__main__":
    app = EC_Toolkit()
    app.mainloop()
    exit()
